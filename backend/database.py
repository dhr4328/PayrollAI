import os
import sqlite3
import openpyxl
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(__file__), "payroll.db")

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Create employees table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS employees (
        emp_code TEXT PRIMARY KEY,
        category TEXT,
        employee_name TEXT,
        unit TEXT,
        floor TEXT,
        department TEXT,
        contractor TEXT,
        doj TEXT,
        bank_name TEXT,
        account_no TEXT,
        ifsc TEXT,
        uan TEXT,
        esic TEXT,
        aadhar TEXT,
        salary_type TEXT,
        per_day_rate REAL,
        fixed_pay REAL
    )
    """)
    
    # Create attendance & payroll records table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS payroll_records (
        emp_code TEXT PRIMARY KEY,
        present REAL,
        hours_ded_hr REAL,
        extra_duty_hrs REAL,
        absent REAL,
        ph REAL,
        weekly_off REAL,
        per_piece REAL,
        paid_days REAL,
        total_days REAL,
        salary REAL,
        hours_ded_amt REAL,
        extra_pay REAL,
        difference_amount REAL,
        bin_card_amount REAL,
        total_earning REAL,
        other_deduction REAL,
        mediclaim_deduction REAL,
        shoes_uniform REAL,
        total_payable_salary REAL,
        ee_pf REAL,
        esi_ee REAL,
        pt REAL,
        er_pf REAL,
        esi_er REAL,
        net_pay REAL,
        remarks TEXT,
        lwf REAL,
        FOREIGN KEY (emp_code) REFERENCES employees (emp_code)
    )
    """)
    
    conn.commit()
    conn.close()
    print("Database tables initialized successfully.")
    
    # Initialize Vector DB
    try:
        import vector_db
        vector_db.init_vector_db()
    except Exception as e:
        print(f"Error initializing VectorDB: {e}")


def clean_float(val):
    if val is None or val == "" or val == "-":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def clean_str(val):
    if val is None or val == "-":
        return ""
    return str(val).strip()

def clean_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = clean_str(val)
    if not val_str:
        return ""
    # Try parsing different formats if needed, or return as is
    return val_str

def load_data_from_excel(excel_path):
    if not os.path.exists(excel_path):
        print(f"Excel file not found at {excel_path}")
        return
        
    print(f"Loading data from {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Clear existing data to avoid conflicts on reload
    cursor.execute("DELETE FROM payroll_records")
    cursor.execute("DELETE FROM employees")
    
    records_count = 0
    # Data rows start from row 5
    for r_idx in range(5, ws.max_row + 1):
        # Read full row
        row = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 47)]
        
        # Check if row is empty or if we reached the total/summary row
        # Row 156 has Col 17 as 'Total'
        per_day_val = row[17]
        emp_code = clean_str(row[3]) # Emp Code2 is at Col 3, Col 1 is formula '=+D...'
        emp_name = clean_str(row[4])
        
        if not emp_code or not emp_name or per_day_val == "Total":
            continue
            
        category = clean_str(row[2])
        unit = clean_str(row[5])
        floor = clean_str(row[6])
        department = clean_str(row[7])
        contractor = clean_str(row[8])
        doj = clean_date(row[9])
        bank_name = clean_str(row[10])
        account_no = clean_str(row[11])
        ifsc = clean_str(row[12])
        uan = clean_str(row[13])
        esic = clean_str(row[14])
        aadhar = clean_str(row[15])
        salary_type = clean_str(row[16])
        per_day_rate = clean_float(row[17])
        fixed_pay = clean_float(row[18])
        
        # Insert into employees
        cursor.execute("""
        INSERT OR REPLACE INTO employees (
            emp_code, category, employee_name, unit, floor, department, contractor, 
            doj, bank_name, account_no, ifsc, uan, esic, aadhar, salary_type, 
            per_day_rate, fixed_pay
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            emp_code, category, emp_name, unit, floor, department, contractor,
            doj, bank_name, account_no, ifsc, uan, esic, aadhar, salary_type,
            per_day_rate, fixed_pay
        ))
        
        # Parse attendance/payroll fields
        present = clean_float(row[19])
        hours_ded_hr = clean_float(row[20])
        extra_duty_hrs = clean_float(row[21])
        absent = clean_float(row[22])
        ph = clean_float(row[23])
        weekly_off = clean_float(row[24])
        per_piece = clean_float(row[25])
        paid_days = clean_float(row[26])
        total_days = clean_float(row[27])
        salary = clean_float(row[28])
        hours_ded_amt = clean_float(row[29])
        extra_pay = clean_float(row[30])
        difference_amount = clean_float(row[31])
        bin_card_amount = clean_float(row[32])
        total_earning = clean_float(row[33])
        other_deduction = clean_float(row[34])
        mediclaim_deduction = clean_float(row[35])
        shoes_uniform = clean_float(row[36])
        total_payable_salary = clean_float(row[37])
        ee_pf = clean_float(row[38])
        esi_ee = clean_float(row[39])
        pt = clean_float(row[40])
        er_pf = clean_float(row[41])
        esi_er = clean_float(row[42])
        net_pay = clean_float(row[43])
        remarks = clean_str(row[44])
        lwf = clean_float(row[45])
        
        # Insert into payroll_records
        cursor.execute("""
        INSERT OR REPLACE INTO payroll_records (
            emp_code, present, hours_ded_hr, extra_duty_hrs, absent, ph, weekly_off,
            per_piece, paid_days, total_days, salary, hours_ded_amt, extra_pay,
            difference_amount, bin_card_amount, total_earning, other_deduction,
            mediclaim_deduction, shoes_uniform, total_payable_salary, ee_pf,
            esi_ee, pt, er_pf, esi_er, net_pay, remarks, lwf
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            emp_code, present, hours_ded_hr, extra_duty_hrs, absent, ph, weekly_off,
            per_piece, paid_days, total_days, salary, hours_ded_amt, extra_pay,
            difference_amount, bin_card_amount, total_earning, other_deduction,
            mediclaim_deduction, shoes_uniform, total_payable_salary, ee_pf,
            esi_ee, pt, er_pf, esi_er, net_pay, remarks, lwf
        ))
        
        records_count += 1
        
    conn.commit()
    conn.close()
    print(f"Excel data loaded into SQLite database successfully. Total records: {records_count}")
    return records_count

# DB Query Helpers for API Routers
def get_all_employees(department=None, category=None):
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    SELECT e.*, p.* 
    FROM employees e
    LEFT JOIN payroll_records p ON e.emp_code = p.emp_code
    WHERE 1=1
    """
    params = []
    if department:
        query += " AND LOWER(e.department) LIKE ?"
        params.append(f"%{department.lower()}%")
    if category:
        query += " AND LOWER(e.category) = ?"
        params.append(category.lower())
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]

def get_employee(emp_code):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT e.*, p.* 
    FROM employees e
    LEFT JOIN payroll_records p ON e.emp_code = p.emp_code
    WHERE e.emp_code = ?
    """, (emp_code,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

def get_attendance_summary():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM payroll_records WHERE present > 0")
    present_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM payroll_records WHERE absent > 0")
    absent_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM payroll_records WHERE extra_duty_hrs > 0")
    overtime_count = cursor.fetchone()[0]
    
    conn.close()
    return {
        "present": present_count,
        "absent": absent_count,
        "overtime_count": overtime_count
    }

def get_absent_employees():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT e.employee_name as name, e.emp_code as empCode, e.department, p.absent
    FROM employees e
    JOIN payroll_records p ON e.emp_code = p.emp_code
    WHERE p.absent > 0
    """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]

def get_overtime_report(department=None):
    conn = get_db_connection()
    cursor = conn.cursor()
    if department:
        cursor.execute("""
        SELECT e.employee_name as name, e.emp_code as empCode, e.department, p.extra_duty_hrs as extraDutyHrs, p.extra_pay as extraPay
        FROM employees e
        JOIN payroll_records p ON e.emp_code = p.emp_code
        WHERE p.extra_duty_hrs > 0 AND LOWER(e.department) LIKE ?
        """, (f"%{department.lower()}%",))
    else:
        cursor.execute("""
        SELECT e.employee_name as name, e.emp_code as empCode, e.department, p.extra_duty_hrs as extraDutyHrs, p.extra_pay as extraPay
        FROM employees e
        JOIN payroll_records p ON e.emp_code = p.emp_code
        WHERE p.extra_duty_hrs > 0
        """)
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]

def get_payroll_summary():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT 
        COUNT(emp_code) as count,
        SUM(total_earning) as totalGross,
        SUM(ee_pf + esi_ee + pt + other_deduction + mediclaim_deduction + shoes_uniform + lwf) as totalDeductions,
        SUM(net_pay) as totalNet,
        SUM(ee_pf + er_pf) as totalPF,
        SUM(esi_ee + esi_er) as totalESI
    FROM payroll_records
    """)
    row = cursor.fetchone()
    conn.close()
    res = dict(row)
    # Handle None values for empty tables
    for k in res:
        if res[k] is None:
            res[k] = 0
    return res

def get_department_summary():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT 
        e.department,
        COUNT(e.emp_code) as count,
        SUM(p.net_pay) as totalNet
    FROM employees e
    JOIN payroll_records p ON e.emp_code = p.emp_code
    GROUP BY e.department
    """)
    rows = cursor.fetchall()
    conn.close()
    return {row["department"]: {"count": row["count"], "totalNet": row["totalNet"]} for row in rows}


def get_record_count() -> int:
    """Return the number of employee records currently in the DB."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM employees")
        count = cursor.fetchone()[0]
        conn.close()
        return count
    except Exception:
        return 0


def load_data_from_dataframe(df) -> int:
    """
    Load payroll data from a pandas DataFrame (produced by the upload router
    after column-mapping). Completely replaces existing data.

    The DataFrame columns should already be renamed to canonical field names
    (emp_code, employee_name, per_day_rate, paid_days, net_pay, ...).
    Any column not in the schema is silently ignored.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # Clear existing data
    cursor.execute("DELETE FROM payroll_records")
    cursor.execute("DELETE FROM employees")

    def _f(row, col, default=0.0):
        """Safe float extraction from a DataFrame row."""
        val = row.get(col, default)
        if val is None or str(val).strip() in ("", "-", "nan"):
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    def _s(row, col, default=""):
        """Safe string extraction from a DataFrame row."""
        val = row.get(col, default)
        if val is None or str(val).strip() in ("-", "nan"):
            return default
        return str(val).strip()

    records_count = 0
    for _, row in df.iterrows():
        row = row.to_dict()
        emp_code = _s(row, "emp_code")
        if not emp_code:
            continue

        # ── Insert employee master ──
        cursor.execute("""
        INSERT OR REPLACE INTO employees (
            emp_code, category, employee_name, unit, floor, department, contractor,
            doj, bank_name, account_no, ifsc, uan, esic, aadhar, salary_type,
            per_day_rate, fixed_pay
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            emp_code,
            _s(row, "category"),
            _s(row, "employee_name"),
            _s(row, "unit"),
            _s(row, "floor"),
            _s(row, "department"),
            _s(row, "contractor"),
            _s(row, "doj"),
            _s(row, "bank_name"),
            _s(row, "account_no"),
            _s(row, "ifsc"),
            _s(row, "uan"),
            _s(row, "esic"),
            _s(row, "aadhar"),
            _s(row, "salary_type"),
            _f(row, "per_day_rate"),
            _f(row, "fixed_pay"),
        ))

        # ── Attendance & Basic Values ──
        present = _f(row, "present")
        hours_ded_hr = _f(row, "hours_ded_hr")
        extra_duty_hrs = _f(row, "extra_duty_hrs")
        absent = _f(row, "absent")
        ph = _f(row, "ph")
        weekly_off = _f(row, "weekly_off")
        per_piece = _f(row, "per_piece")
        
        total_days = _f(row, "total_days")
        if total_days == 0:
            total_days = (present + absent + ph + weekly_off) or 28.0

        paid_days = _f(row, "paid_days")
        if paid_days == 0 and (present > 0 or ph > 0 or weekly_off > 0):
            paid_days = present + ph + weekly_off

        per_day_rate = _f(row, "per_day_rate")
        fixed_pay = _f(row, "fixed_pay")

        # ── Earnings Calculation ──
        salary = _f(row, "salary")
        if salary == 0 and (per_day_rate > 0 or fixed_pay > 0):
            if per_day_rate > 0:
                salary = round(per_day_rate * paid_days, 2)
            else:
                salary = fixed_pay

        hours_ded_amt = _f(row, "hours_ded_amt")
        if hours_ded_amt == 0 and hours_ded_hr > 0 and per_day_rate > 0:
            hours_ded_amt = round((per_day_rate / 8.0) * hours_ded_hr, 2)

        extra_pay = _f(row, "extra_pay")
        if extra_pay == 0 and extra_duty_hrs > 0 and per_day_rate > 0:
            extra_pay = round((per_day_rate / 8.0) * extra_duty_hrs, 2)

        difference_amount = _f(row, "difference_amount")
        bin_card_amount = _f(row, "bin_card_amount")

        total_earning = _f(row, "total_earning")
        if total_earning == 0:
            total_earning = round(salary + extra_pay + bin_card_amount + difference_amount - hours_ded_amt, 2)

        # ── Deductions & Statutory ──
        other_deduction = _f(row, "other_deduction")
        mediclaim_deduction = _f(row, "mediclaim_deduction")
        shoes_uniform = _f(row, "shoes_uniform")

        total_payable_salary = _f(row, "total_payable_salary")
        if total_payable_salary == 0 and total_earning > 0:
            total_payable_salary = round(total_earning - other_deduction - mediclaim_deduction - shoes_uniform, 2)

        ee_pf = _f(row, "ee_pf")
        if ee_pf == 0 and salary > 0:
            pf_base = min(15000.0, salary)
            ee_pf = round(pf_base * 0.12, 2)

        esi_ee = _f(row, "esi_ee")
        if esi_ee == 0 and total_earning > 0 and total_earning <= 21000:
            esi_ee = round(total_earning * 0.0075, 2)

        er_pf = _f(row, "er_pf")
        if er_pf == 0 and salary > 0:
            pf_base = min(15000.0, salary)
            er_pf = round(pf_base * 0.13, 2)

        esi_er = _f(row, "esi_er")
        if esi_er == 0 and total_earning > 0 and total_earning <= 21000:
            esi_er = round(total_earning * 0.0325, 2)

        pt = _f(row, "pt")
        lwf = _f(row, "lwf")

        net_pay = _f(row, "net_pay")
        if net_pay == 0 and total_earning > 0:
            net_pay = round(total_payable_salary - ee_pf - esi_ee - pt - lwf, 2)

        remarks = _s(row, "remarks")

        # ── Insert payroll record ──
        cursor.execute("""
        INSERT OR REPLACE INTO payroll_records (
            emp_code, present, hours_ded_hr, extra_duty_hrs, absent, ph, weekly_off,
            per_piece, paid_days, total_days, salary, hours_ded_amt, extra_pay,
            difference_amount, bin_card_amount, total_earning, other_deduction,
            mediclaim_deduction, shoes_uniform, total_payable_salary, ee_pf,
            esi_ee, pt, er_pf, esi_er, net_pay, remarks, lwf
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            emp_code, present, hours_ded_hr, extra_duty_hrs, absent, ph, weekly_off,
            per_piece, paid_days, total_days, salary, hours_ded_amt, extra_pay,
            difference_amount, bin_card_amount, total_earning, other_deduction,
            mediclaim_deduction, shoes_uniform, total_payable_salary, ee_pf,
            esi_ee, pt, er_pf, esi_er, net_pay, remarks, lwf
        ))
        records_count += 1

    conn.commit()
    conn.close()
    print(f"DataFrame import complete. Total records: {records_count}")
    return records_count


def update_employee_record(emp_code: str, updates: dict) -> bool:
    """
    Update employee master or payroll record fields dynamically, then recalculate totals.
    updates can contain: per_day_rate, present, absent, extra_duty_hrs, other_deduction, department, etc.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    emp_code = emp_code.upper().strip()

    emp_fields = {"category", "employee_name", "unit", "floor", "department", "contractor",
                  "bank_name", "account_no", "ifsc", "uan", "esic", "aadhar", "salary_type",
                  "per_day_rate", "fixed_pay"}
                  
    payroll_fields = {"present", "hours_ded_hr", "extra_duty_hrs", "absent", "ph", "weekly_off",
                      "per_piece", "other_deduction", "mediclaim_deduction", "shoes_uniform", "remarks"}

    emp_updates = {k: v for k, v in updates.items() if k in emp_fields}
    pay_updates = {k: v for k, v in updates.items() if k in payroll_fields}

    if emp_updates:
        set_clause = ", ".join(f"{k} = ?" for k in emp_updates.keys())
        params = list(emp_updates.values()) + [emp_code]
        cursor.execute(f"UPDATE employees SET {set_clause} WHERE emp_code = ?", params)

    if pay_updates:
        set_clause = ", ".join(f"{k} = ?" for k in pay_updates.keys())
        params = list(pay_updates.values()) + [emp_code]
        cursor.execute(f"UPDATE payroll_records SET {set_clause} WHERE emp_code = ?", params)

    conn.commit()
    conn.close()

    recalculate_employee_payroll(emp_code)
    return True


def update_all_employees_records(updates: dict) -> int:
    """
    Update employee master or payroll record fields dynamically for ALL employees in the database,
    and recalculate payroll for every employee in the organization.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    emp_fields = {"category", "employee_name", "unit", "floor", "department", "contractor",
                  "bank_name", "account_no", "ifsc", "uan", "esic", "aadhar", "salary_type",
                  "per_day_rate", "fixed_pay"}
                  
    payroll_fields = {"present", "hours_ded_hr", "extra_duty_hrs", "absent", "ph", "weekly_off",
                      "per_piece", "other_deduction", "mediclaim_deduction", "shoes_uniform", "remarks"}

    emp_updates = {k: v for k, v in updates.items() if k in emp_fields}
    pay_updates = {k: v for k, v in updates.items() if k in payroll_fields}

    if emp_updates:
        set_clause = ", ".join(f"{k} = ?" for k in emp_updates.keys())
        params = list(emp_updates.values())
        cursor.execute(f"UPDATE employees SET {set_clause}", params)

    if pay_updates:
        set_clause = ", ".join(f"{k} = ?" for k in pay_updates.keys())
        params = list(pay_updates.values())
        cursor.execute(f"UPDATE payroll_records SET {set_clause}", params)

    conn.commit()

    cursor.execute("SELECT emp_code FROM employees")
    rows = cursor.fetchall()
    emp_codes = [r["emp_code"] for r in rows]
    conn.close()

    for code in emp_codes:
        recalculate_employee_payroll(code)

    return len(emp_codes)



def recalculate_employee_payroll(emp_code: str):
    """Recalculate salary, extra pay, total earning, PF, ESI, and net pay for an employee."""
    conn = get_db_connection()
    cursor = conn.cursor()
    emp_code = emp_code.upper().strip()

    cursor.execute("""
    SELECT e.per_day_rate, e.fixed_pay, p.* 
    FROM employees e
    JOIN payroll_records p ON e.emp_code = p.emp_code
    WHERE e.emp_code = ?
    """, (emp_code,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return

    r = dict(row)
    per_day_rate = float(r.get("per_day_rate") or 0)
    fixed_pay = float(r.get("fixed_pay") or 0)
    present = float(r.get("present") or 0)
    ph = float(r.get("ph") or 0)
    weekly_off = float(r.get("weekly_off") or 0)
    absent = float(r.get("absent") or 0)
    extra_duty_hrs = float(r.get("extra_duty_hrs") or 0)
    hours_ded_hr = float(r.get("hours_ded_hr") or 0)

    paid_days = present + ph + weekly_off
    total_days = (present + absent + ph + weekly_off) or 28.0

    if per_day_rate > 0:
        salary = round(per_day_rate * paid_days, 2)
    else:
        salary = fixed_pay

    extra_pay = round((per_day_rate / 8.0) * extra_duty_hrs, 2) if (extra_duty_hrs > 0 and per_day_rate > 0) else float(r.get("extra_pay") or 0)
    hours_ded_amt = round((per_day_rate / 8.0) * hours_ded_hr, 2) if (hours_ded_hr > 0 and per_day_rate > 0) else float(r.get("hours_ded_amt") or 0)

    diff_amt = float(r.get("difference_amount") or 0)
    bin_card = float(r.get("bin_card_amount") or 0)

    total_earning = round(salary + extra_pay + bin_card + diff_amt - hours_ded_amt, 2)

    other_ded = float(r.get("other_deduction") or 0)
    mediclaim = float(r.get("mediclaim_deduction") or 0)
    shoes = float(r.get("shoes_uniform") or 0)

    total_payable = round(total_earning - other_ded - mediclaim - shoes, 2)

    pf_base = min(15000.0, salary)
    ee_pf = round(pf_base * 0.12, 2) if salary > 0 else 0.0
    er_pf = round(pf_base * 0.13, 2) if salary > 0 else 0.0

    esi_ee = round(total_earning * 0.0075, 2) if (total_earning > 0 and total_earning <= 21000) else 0.0
    esi_er = round(total_earning * 0.0325, 2) if (total_earning > 0 and total_earning <= 21000) else 0.0

    pt = 200.0 if total_payable > 12000 else 0.0
    lwf = float(r.get("lwf") or 0)

    net_pay = round(total_payable - ee_pf - esi_ee - pt - lwf, 2)

    cursor.execute("""
    UPDATE payroll_records SET
        paid_days = ?, total_days = ?, salary = ?, hours_ded_amt = ?, extra_pay = ?,
        total_earning = ?, total_payable_salary = ?, ee_pf = ?, esi_ee = ?,
        er_pf = ?, esi_er = ?, pt = ?, net_pay = ?
    WHERE emp_code = ?
    """, (
        paid_days, total_days, salary, hours_ded_amt, extra_pay,
        total_earning, total_payable, ee_pf, esi_ee,
        er_pf, esi_er, pt, net_pay, emp_code
    ))
    conn.commit()
    conn.close()


